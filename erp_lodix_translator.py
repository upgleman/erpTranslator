#-*- coding:utf-8 -*-

import os
import sys
import json
import time
import math
import shutil
import numpy as np
import pandas as pd

from openpyxl import load_workbook


EXCEL_HEADER = [
        '물류센터명', '운송타입', '배송지명', '자체관리코드', '배송지 주소',
        '배송지 상세주소', '위도(Y)', '경도(X)', '주문유형','운송부피(CBM)',
        '운송중량(kg)', '박스수량', '물품가액(원)', '상품분류', '상품명',
        '상품수량', '개별중량(kg)', '개별원가(원)', '하차유형', '하차소요시간',
        '지정배송시간', '배송가능시간(시작)', '배송가능시간(종료)', '배송회피시간', '수화인명',
        '수화인연락처', '배송지특이사항', '배송차량', '배송순서'
    ]
    # '물품가액(원)','하차소요시간','수화인명','수화인연락처','배송지특이사항'

ERP_HEADER = [
        'CENTER_NM','ORDER_TYPE','LOCATION_NM', 'LOC_CUSTOM_CD', 'ADDRESS',
        'SUB_ADDRESS','Y','X','ORDER_CLASS','ORDER_VOLUME',
        'ORDER_WEIGHT','BOX_NUM', '','ITEM_TYPE','ITEM_NM',
        'ITEM_COUNT','ITEM_WEIGHT','ITEM_COST','UNLOADING_TYPE','',
        'ORDER_TIME','S_ORDER_TIME','E_ORDER_TIME','FORBIDDEN_TIME','',
        '', '', 'OLD_CAR_NUM','OLD_VISIT_ORDER'
    ]

REDUCED_EXCEL_HEADER_FOR_V1 = [
        '물류센터명', '운송타입', '배송지명', '배송지 주소', '배송지 상세주소',
        '주문유형','위도(Y)','경도(X)','운송부피(CBM)','운송중량(kg)',
        '박스수량','상품분류','상품명','상품수량','개별중량(kg)',
        '개별원가(원)','자체관리코드','하차유형','지정배송시간','배송가능시간(시작)',
        '배송가능시간(종료)','배송회피시간','배송차량', '배송순서'
    ]

EXCEL_HEADER_V1 = ['물류센터명','운송타입','배송지명','배송지 주소','배송지 상세주소',
                '주문유형','위도(Y)','경도(X)','운송부피(CBM)','운송중량(kg)',
                '박스수량','상품분류','상품명','상품수량','개별중량(kg)',
                '개별원가','자체관리코드','하차유형','점착시간','점착시작시간',
                '점착종료시간','회피시간','배송차량','배송순서']

ERP_HEADER_V1 = ['CENTER_NM','ORDER_TYPE','LOCATION_NM','ADDRESS','SUB_ADDRESS',
                'ORDER_CLASS','Y','X','ORDER_VOLUME','ORDER_WEIGHT',
                'BOX_NUM','ITEM_TYPE','ITEM_NM','ITEM_COUNT','ITEM_WEIGHT',
                'ITEM_COST','LOC_CUSTOM_CD','UNLOADING_TYPE','ORDER_TIME','S_ORDER_TIME',
                'E_ORDER_TIME','FORBIDDEN_TIME','OLD_CAR_NUM','OLD_VISIT_ORDER']

CUSTOM_VALUE_WITH_HEADER = {
    '운송타입': {
        'nan': '일반',
        '0': '일반',
        '1': '회수'
    },
    '주문유형': {
        '0': '일반',
        '1': '냉장'
    },
    '하차유형': {
        '0': '수작업',
        '1': '지게차'
    }
}

VAL_RANGE_WITH_HEADER = {
    # 대한민국 지도에 관한 일반정보의 경도범위는 124 – 132, 위도범위는 33 – 43 이다.
    '위도(Y)': [20.0, 55.0],
    '경도(X)': [110.0, 140.0]
}

COLS_TO_EMPTY = ['운송부피(CBM)','운송중량(kg)','박스수량','배송회피시간', '하차유형', '물품가액(원)']

COLS_TO_REMOVE_ZERO = ['개별중량(kg)','상품수량', '개별원가(원)','지정배송시간','배송가능시간(시작)','배송가능시간(종료)']

COLS_TO_REMOVE_ABNORMAL_HISTORY = ['배송차량', '배송순서']

COLS_TO_REMOVE_QUOTATION_MARK = ['배송지 주소', '배송지 상세주소']


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def integrate_kt_and_mns_order(_kt_file_path, _mns_file_path, _kt_quotechar="'", _mns_quotechar='"'):
    # KT and MNS integration part


    kt_order_df = pd.read_csv(_kt_file_path, quotechar = _kt_quotechar, dtype = 'str')

    mns_order_df = pd.read_csv(_mns_file_path, quotechar = _mns_quotechar, dtype = 'str')

    # Remove comma in address
    kt_address_list = kt_order_df['ADDRESS'].tolist()
    kt_sub_address_list = kt_order_df['SUB_ADDRESS'].tolist()

    for idx, elem in enumerate(kt_address_list):
        elem = elem.replace(' ,', ' ')
        elem = elem.replace(', ', ' ')
        elem = elem.replace(',', ' ')
        elem = elem.replace('"','')
        kt_address_list[idx] = elem.replace("'",'')
        # print(elem)
    for idx, elem in enumerate(kt_sub_address_list):
        elem = elem.replace(' ,', ' ')
        elem = elem.replace(', ', ' ')
        elem = elem.replace(',', ' ')
        elem = elem.replace('"','')
        kt_sub_address_list[idx] = elem.replace("'",'')
        # print(elem)

    mns_address_list = mns_order_df['ADDRESS'].tolist()
    mns_sub_address_list = mns_order_df['SUB_ADDRESS'].tolist()

    for idx, elem in enumerate(mns_address_list):
        elem = elem.replace(' ,', ' ')
        elem = elem.replace(', ', ' ')
        elem = elem.replace(',', ' ')
        elem = elem.replace('"','')
        mns_address_list[idx] = elem.replace("'",'')
        # print(elem)
    for idx, elem in enumerate(mns_sub_address_list):
        elem = elem.replace(' ,', ' ')
        elem = elem.replace(', ', ' ')
        elem = elem.replace(',', ' ')
        elem = elem.replace('"','')
        mns_sub_address_list[idx] = elem.replace("",'')
        # print(elem)

    kt_order_df['ADDRESS'] = kt_address_list
    kt_order_df['SUB_ADDRESS'] = kt_sub_address_list
    mns_order_df['ADDRESS'] = mns_address_list
    mns_order_df['SUB_ADDRESS'] = mns_sub_address_list

    # Set time info as 0 - order time, forbidden time
    kt_order_df[["ORDER_TIME","S_ORDER_TIME","E_ORDER_TIME","FORBIDDEN_TIME"]] = 0
    mns_order_df[["ORDER_TIME","S_ORDER_TIME","E_ORDER_TIME","FORBIDDEN_TIME"]] = 0

    # Get and align column order
    kt_order_cols = kt_order_df.columns
    mns_order_cols = mns_order_df.columns

    mns_order_df = mns_order_df[kt_order_cols]

    # Concat two order dfs - KT and MNS
    concatenated_order_df = pd.concat([kt_order_df, mns_order_df]).reset_index()
    result_order_df = concatenated_order_df.copy(deep=True)

    return [result_order_df, [kt_order_df, mns_order_df]]


def apply_rules_to_integrated_order(_integrated_df, _rule_json_file_path = './resources/correction_rules.json'):
    correction_result_dict = {}
    correction_result_list = []

    result_df = _integrated_df.copy(deep=True)

    with open('./resources/correction_rules.json') as f:
        correction_rule_data = json.load(f)
        rule_list = correction_rule_data['rules']
    
    for idx, rrow in result_df.iterrows():
        cols_to_correct = {}

        for tmp_rule_json in rule_list:
            condition_list = tmp_rule_json['condition']
            action_list = tmp_rule_json['action']

            need_to_correct = True
            for key, val in condition_list.items():
                if rrow[key] != val:
                    need_to_correct = False
                    break
            
            if need_to_correct:
                for key, val in action_list.items():
                    cols_to_correct[key] = val
        
        if len(cols_to_correct.keys()) > 0:
            correction_result_list.append([idx, cols_to_correct])
        
    for elem in correction_result_list:
        # print(elem[0], elem[1])

        for key, val in elem[1].items():
            result_df.loc[elem[0], key] = val

    return result_df





def convert_etl_format_to_excel_format(_etl_df):
    excel_converted_df = None
    tmp_list_size = len(_etl_df.index)

    print(_etl_df.iloc[0])

    for col_idx in range(0, len(EXCEL_HEADER)):
        print(col_idx, EXCEL_HEADER[col_idx], ERP_HEADER[col_idx])
        col_to_pick = ERP_HEADER[col_idx]
        col_to_insert = EXCEL_HEADER[col_idx]
        
        if col_to_pick == "" or col_to_pick == '' or len(col_to_pick) == 0:
            list_to_insert = [''] * tmp_list_size
        else:
            list_to_insert = _etl_df[col_to_pick].tolist()

        if col_to_insert in CUSTOM_VALUE_WITH_HEADER.keys():
            tmp_val_dict = CUSTOM_VALUE_WITH_HEADER[col_to_insert]
            for idx, elem in enumerate(list_to_insert):
                # print(col_to_insert, idx, elem, type(elem), tmp_val_dict.keys())
                if type(elem) == int:
                    print(str(elem), tmp_val_dict.keys())
                    if str(elem) in tmp_val_dict.keys():
                        list_to_insert[idx] = tmp_val_dict[str(list_to_insert[idx])]
                elif type(elem) == float:
                    if elem.isnull():
                        elem = 'nan'
                    else:
                        elem = str(int(elem))
                    
                    if str(elem) in tmp_val_dict.keys():
                        list_to_insert[idx] = tmp_val_dict[str(list_to_insert[idx])]

        
        if col_to_insert in COLS_TO_EMPTY:
            list_to_insert = [''] * tmp_list_size

        if col_to_insert in VAL_RANGE_WITH_HEADER.keys():
            min_val = float(min(list(VAL_RANGE_WITH_HEADER[col_to_insert])))
            max_val = float(max(list(VAL_RANGE_WITH_HEADER[col_to_insert])))

            for idx, elem in enumerate(list_to_insert):
                try:
                    tmp_val = float(elem)
                except ValueError:
                    # Handle the exception
                    list_to_insert[idx] = ''
                    continue

                if math.isnan(tmp_val):
                    continue

                if tmp_val > max_val or tmp_val < min_val:
                    list_to_insert[idx] = ''
                else:
                    pass
        
        if col_to_insert in COLS_TO_REMOVE_ZERO:
            for idx, elem in enumerate(list_to_insert):
                if elem == 0 or elem == '0' or elem == '0.00':
                    list_to_insert[idx] = ''
        
        if col_to_insert in COLS_TO_REMOVE_QUOTATION_MARK:
            for idx, elem in enumerate(list_to_insert):
                elem = elem.replace('"', '')
                elem = elem.replace("'", '')
                list_to_insert[idx] = elem

        if excel_converted_df is None:
            excel_converted_df = pd.DataFrame({col_to_insert: list_to_insert})
            pass
        else:
            #print(col_to_insert)
            excel_converted_df[col_to_insert] = list_to_insert

    list_old_vehicle_num = _etl_df['OLD_CAR_NUM'].tolist()
    list_old_vehicle_visit_idx = _etl_df['OLD_VISIT_ORDER'].tolist()

    for idx in range(0, len(list_old_vehicle_visit_idx)):
        if list_old_vehicle_visit_idx[idx] == '0':
            if list_old_vehicle_num[idx] == '0':
                list_old_vehicle_num[idx] = ''
                list_old_vehicle_visit_idx[idx] = ''
    
    excel_converted_df['배송차량'] = list_old_vehicle_num
    excel_converted_df['배송순서'] = list_old_vehicle_visit_idx

    for idx in range(0, len(list_old_vehicle_visit_idx)):
        if list_old_vehicle_visit_idx[idx] == '0':
            list_old_vehicle_visit_idx[idx] = ''

    excel_converted_df = excel_converted_df[EXCEL_HEADER]

    return excel_converted_df


def compare_two_dfs(df1, df2):
    df1_row_size = len(df1.index)
    df2_row_size = len(df2.index)

    df1_columns = df1.columns
    df2_columns = df2.columns

    if df1_row_size != df2_row_size:
        # print("SIZE IS DIFFERENT!!!")
        return False
    
    df1_columns = df1_columns[:-4]
    
    for i in range(0, len(df1_columns)):
        if df1_columns[i] != df2_columns[i]:
            # print("COLUMN ORDER IS DIFFERENT!!!")
            return False
    
    for i in range(0, df1_row_size):
        diff_list = []
        for col_name in df1_columns:
            if df1.loc[i, col_name] != df2.loc[i, col_name]:
                diff_list.append([col_name, df1.loc[i, col_name], df2.loc[i, col_name]])
    
        # if len(diff_list) > 0:
        #     print(i, end=": ")
        #     for elem in diff_list:
        #         print(elem, end=",")
        #     print()

    
def generate_excel_for_web_upload(_df_to_write, _result_file_path, _template_version="v2"):
    # 버전별로 column 변경하고, column명도 적용하기


    if _template_version == "v2":
        # column 추리고 순서와 이름 바꾸기
        # V2는 그대로 유지

        # 엑셀에 쓰기
        write_order_to_excel(_df_to_write, _result_file_path, _template_file_path='./template/delivery_order_template_v2.linkus', _start_row=8)
    elif _template_version == "v1":
        # column 추리고 순서와 이름 바꾸기
        df_to_write = _df_to_write[REDUCED_EXCEL_HEADER_FOR_V1]
        df_to_write.columns = EXCEL_HEADER_V1

        # 엑셀에 쓰기
        write_order_to_excel(_df_to_write, _result_file_path, _template_file_path='./template/delivery_order_template_v1.linkus', _start_row=7)
    pass

def write_order_to_excel(_df_to_write, _result_file_path, _template_file_path='./template/delivery_order_template.linkus', _start_row=8):
    # Get excel template
    

    # Align columns to fit excel format (Result: 'converted_etl_df')


    # Copy template and write to copied template (=result file)
    abs_template_file_path = resource_path(_template_file_path)
    shutil.copy(abs_template_file_path, _result_file_path)

    if os.name == "nt":
        writer = pd.ExcelWriter(_result_file_path, engine='openpyxl', mode='a')
        writer.book = load_workbook(_result_file_path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    else:
        writer = pd.ExcelWriter(_result_file_path, engine='openpyxl', mode='w')
        writer.book = load_workbook(_result_file_path)
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)

    _df_to_write.to_excel(writer, sheet_name='운송주문입력', startrow=_start_row, startcol=0, header=False, index=False, )
    writer.save()
    writer.close()

#./home/atc/code/ETL_ERP_LODIX_translator/order_data/20210311/1031_01_20210311_KT.CSV

if __name__ == "__main__":

    mid_result_dir = './mid_result/'

    center_id = 1031
    cluster_id = 1
    date = 20210327

    separated_file_dir = './order_data/{0}/raw/'.format(date)

    for i in [1, 2, 3, 4, 5, 6]:
        print("Processing order data of cluster {0}".format(i))

        kt_file_name = '{0}_{1:02d}_{2}_KT.csv'.format(center_id, i, date)
        mns_file_name = '{0}_{1:02d}_{2}_MnS.csv'.format(center_id, i, date)
        integrated_file_name = '{0}_{1:02d}_{2}_integrated.csv'.format(center_id, i, date)

        result_dir = './result/{0}/'.format(date)

        #kt_file_name = '1031_08_20210309_KT.csv'
        #mns_file_name = '1031_08_20210309_MnS.csv'
        #integrated_file_name = '1031_08_20210309_integrated.csv'

        kt_file_path = separated_file_dir + kt_file_name
        mns_file_path = separated_file_dir + mns_file_name

        integrated_file_path = result_dir + integrated_file_name

        #kt_file_path = './order_data/20210311/raw/1031_01_20210311_KT.CSV'
        #mns_file_path = './order_data/20210311/raw/1031_01_20210311_MnS.CSV'

        # Get file path as single string - after, use QT

        kt_quotechar = "'"
        mns_quotechar = "'"

        integrated_df, _ = integrate_kt_and_mns_order(kt_file_path, mns_file_path, kt_quotechar, mns_quotechar)

        #concatenated_order_df = apply_rules_to_integrated_order(integrated_df)

        #compare_two_dfs(integrated_df, concatenated_order_df)

        # compare two dataframes


        #concatenated_order_df = pd.read_csv('./1031_01_20210312.csv')
        #integrated_file_name = '1031_01_20210312_integrated.csv'

        # Save to excel
        excel_df = convert_etl_format_to_excel_format(integrated_df)
        #excel_df = convert_etl_format_to_excel_format(concatenated_order_df)

        write_order_to_excel(excel_df, '{0}.xlsx'.format(integrated_file_path[:-4]))

        print("Get integrated order data of cluster {0}".format(i))

    print("FINISH!!!!!!")